import json

from openpyxl import Workbook, load_workbook


def Inputs(data, data2=None):
    if data == '' or data == []:
        return
    if isinstance(data, str) and data2 is None:
        return data
    elif isinstance(data, list):
        if '其他' in data:
            data.remove('其他')
            data.append('其他：%s' % data2)
            return ','.join(data)
        else:
            return ','.join(data)
    if data2 and data2 != '':
        return ','.join([data, data2])


def Basic():
    filename = 'state.json'
    with open(filename, 'r', encoding='utf-8') as f:
        state = json.load(f)
    wb = load_workbook(r'C:\Users\78170\Desktop\数据汇总\1基本情况.xlsx')
    sheet_list = wb.sheetnames
    ws = wb[sheet_list[0]]

    ws.cell(3, 1).value = Inputs(state['name'])
    ws.cell(3, 2).value = Inputs(state['place'])

    ws.cell(3, 3).value = Inputs((state['createDate'].split('T'))[0])
    ws.cell(3, 4).value = Inputs(state['isIndependLegal_isTrue'], state['isIndependLegal_msg'])

    ws.cell(3, 5).value = Inputs(state['museumType_type'], state['museumType_msg'])
    # ws.cell(3, 2).value = '-'.join(state['place'])
    # ws.cell(3, 2).value = '-'.join(state['place'])
    # ws.cell(3, 2).value = '-'.join(state['place'])
    # ws.cell(3, 2).value = '-'.join(state['place'])
    # ws.cell(3, 2).value = '-'.join(state['place'])
    # ws.cell(3, 2).value = '-'.join(state['place'])
    # ws.cell(3, 2).value = '-'.join(state['place'])






    wb.save(r'C:\Users\78170\Desktop\数据汇总\1基本情况.xlsx')

if __name__ == '__main__':
    Basic()